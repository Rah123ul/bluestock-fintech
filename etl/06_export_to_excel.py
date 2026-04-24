import os
import sys
import pandas as pd
from dotenv import load_dotenv
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ── 0. Load env & create engine ──────────────────────────────────────────────
load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

if not all([DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD]):
    # Fallback to DATABASE_URL if individual vars aren't set
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        print("❌ Database credentials not fully set in .env")
        sys.exit(1)
else:
    # Handle railway specific postgres:// -> postgresql://
    DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

engine = create_engine(DATABASE_URL)

OUTPUT_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "bluestock_analysis.xlsx")

# Ensure data directory exists
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

# ── 1. Fetch Data ────────────────────────────────────────────────────────────
print("📥 Fetching data from PostgreSQL...")

query_health = """
    SELECT company_name, overall_score, health_label,
           score_profitability, score_growth, score_leverage, score_cashflow
    FROM fact_ml_scores
    ORDER BY overall_score DESC
"""
df_health = pd.read_sql(query_health, engine)

query_revenue = """
    WITH ranked AS (
        SELECT company_id, year_label, fiscal_year, sales, net_profit, opm_percentage,
               ROW_NUMBER() OVER(PARTITION BY company_id ORDER BY fiscal_year DESC) as rn
        FROM fact_profit_loss
    )
    SELECT c.company_id, c.company_name, r.year_label, r.sales, r.net_profit, r.opm_percentage
    FROM ranked r
    JOIN dim_company c ON r.company_id = c.company_id
    WHERE r.rn <= 5
    ORDER BY c.company_name, r.fiscal_year DESC
"""
df_revenue = pd.read_sql(query_revenue, engine)

query_balance = """
    WITH ranked AS (
        SELECT company_id, year_label, fiscal_year, borrowings, total_assets, debt_to_equity,
               ROW_NUMBER() OVER(PARTITION BY company_id ORDER BY fiscal_year DESC) as rn
        FROM fact_balance_sheet
    )
    SELECT c.company_id, c.company_name, r.year_label, r.borrowings, r.total_assets, r.debt_to_equity
    FROM ranked r
    JOIN dim_company c ON r.company_id = c.company_id
    WHERE r.rn <= 5
    ORDER BY c.company_name, r.fiscal_year DESC
"""
df_balance = pd.read_sql(query_balance, engine)

query_top10 = """
    SELECT *
    FROM fact_ml_scores
    ORDER BY overall_score DESC
    LIMIT 10
"""
df_top10 = pd.read_sql(query_top10, engine)

# ── 2. Write to Excel ────────────────────────────────────────────────────────
print(f"💾 Writing data to {OUTPUT_FILE}...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df_health.to_excel(writer, sheet_name="Health Scores", index=False)
    df_revenue.to_excel(writer, sheet_name="Revenue Analysis", index=False)
    df_balance.to_excel(writer, sheet_name="Balance Sheet", index=False)
    df_top10.to_excel(writer, sheet_name="Top 10 Companies", index=False)

# ── 3. Apply Formatting ──────────────────────────────────────────────────────
print("🎨 Applying Excel formatting...")

wb = load_workbook(OUTPUT_FILE)

# Color mapping for health_label
# EXCELLENT=green, GOOD=blue, AVERAGE=orange, WEAK=red, POOR=darkred
COLOR_MAP = {
    "EXCELLENT": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), # Green
    "GOOD": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),      # Blue
    "AVERAGE": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),   # Orange
    "WEAK": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),      # Red
    "POOR": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")       # Dark Red
}

header_font = Font(bold=True)
white_font = Font(color="FFFFFF")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Bold headers
    for cell in ws[1]:
        cell.font = header_font
        
    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value:
                    # add a little padding
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Apply colors based on health_label if column exists
    # Find the health_label column index
    health_col_idx = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == "health_label":
            health_col_idx = idx + 1
            break
            
    if health_col_idx:
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=health_col_idx).value
            if cell_val in COLOR_MAP:
                fill = COLOR_MAP[cell_val]
                
                # Apply to entire row
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = fill
                    # Make text white for darker backgrounds for readability
                    if cell_val in ["POOR", "WEAK", "GOOD", "AVERAGE", "EXCELLENT"]:
                        # Just apply to POOR and WEAK as they are darkest
                        if cell_val in ["POOR", "WEAK"]:
                            cell.font = white_font

wb.save(OUTPUT_FILE)
print("✅ Excel file generated successfully with formatting!")
